from tkinter import *
from tkcalendar import Calendar, DateEntry
from functools import partial
import pandas as pd
import openpyxl
import os
import dateutil.parser
import datetime
import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
# from PyInquirer import prompt
import pprint
import re
from tkinter import messagebox
from tkinter.font import Font

def stringsandyear_topath(stringsandyear,year):
    string = ""
    for x in stringsandyear:
        print(x)
        if "year" not in x:
            string += x
        else:
            string += f"{year}"
    return string

def stringsandinvoicenumber_topath(stringsandinvoicenumber, invoicenumber, clientname, date):
    string = ""
    for x in stringsandinvoicenumber:
        print(x)
        if ("invoicenumber" in x) :
            string += invoicenumber
        if ("clientname" in x) :
            string += clientname
        if ("date" in x):
            string += date
        if ("date" not in x) and ("clientname" not in x) and ("invoicenumber" not in x) :
            string += x
    print(string)
    return string
def check_invoice_archive(year_of_invoice,outputdir_path,archive_which_invoices_path,invoice_achive_template_path,invoicenumber_pattern):
    print(f"Year to link this invoice to: {year_of_invoice}")
    if not os.path.exists(outputdir_path):
        os.mkdir(outputdir_path)

    if not os.path.exists(archive_which_invoices_path):
        print(f"Because there was no Archive file of the year create one at {archive_which_invoices_path}")
        wb = openpyxl.load_workbook(invoice_achive_template_path)
        # Select the worksheet
        sheet = wb["Tabelle1"]
        # Modify the cell
        sheet["A1"] = f"Rechnungen {year_of_invoice}"
        wb.save(archive_which_invoices_path)
        lastinvoice_year_num = f"{year_of_invoice}-001"
        lastinvoice_year = year_of_invoice
        lastinvoice_num = 1

    invoicenumbers = pd.read_excel(archive_which_invoices_path).iloc[:-2,0] # get the invoice numbers out of the invoice archive
    # search for the first invoicenumer which fits the pattern
    invoicenumber_pattern = invoicenumber_pattern
    lastinvoice_num = 0
    for index, entry in invoicenumbers.iloc[::-1].items():
        if not pd.isnull(entry):
            if re.match(invoicenumber_pattern, entry):
                lastinvoice_year_num = re.match(invoicenumber_pattern, entry)[0]
                lastinvoice_year = int(re.match(invoicenumber_pattern, entry)[1])
                lastinvoice_num = int(re.match(invoicenumber_pattern, entry)[2])
                break
    return lastinvoice_num

def question_next_invoice_number(invoiceyear,lastinvoice_num,invoicenumber_pattern,invoicenumber_pattern_names):
    #get which invoicenumber
    answer1 = "Nimm einfach die Nächste in der Reihe"
    answer2 = "Ich möchte sie selber eingeben"
    invoicenumberquestion_choices = [answer1, answer2]
    thisinvoicenumber = ""
    while not thisinvoicenumber:
        last_inv_numb_str_sugg = ""
        this_inv_numb_str_sugg = ""
        for x in invoicenumber_pattern_names:
            print(x)
            if x in "year":
                last_inv_numb_str_sugg += f"{invoiceyear}"
                this_inv_numb_str_sugg += f"{invoiceyear}"
            elif x in "invoicenumber":
                last_inv_numb_str_sugg += f"{(lastinvoice_num):03}"
                this_inv_numb_str_sugg += f"{(lastinvoice_num + 1):03}"
            else:
                last_inv_numb_str_sugg += x
                this_inv_numb_str_sugg += x


        result = ask_right_invoicenumber(f"Die letzte Rechnungsnummer war {last_inv_numb_str_sugg}. \nSomit wäre die nächste Rechnungsnummer {this_inv_numb_str_sugg}.")

        if result:

            thisinvoicenumber = this_inv_numb_str_sugg
            print("Das ist die Rechnungsnummer: " + thisinvoicenumber)
        if not result:
            root = tk.Tk()
            # withdraw() will make the parent window disappear.
            root.withdraw()
            # shows a dialogue with a string input field
            thisinvoicenumber = tk.simpledialog.askstring('Rechnungsnummer',
                                                       f"Dann kannst du sie jetzt selber eingeben (in dem Format z.b. {last_inv_numb_str_sugg}):",
                                                       parent=root)
            root.destroy()
            if not thisinvoicenumber:
                continue

            while not re.match(invoicenumber_pattern,thisinvoicenumber):
                root = tk.Tk()
                change_place_of_window(root)
                # withdraw() will make the parent window disappear.
                root.withdraw()
                # shows a dialogue with a string input field
                thisinvoicenumber = tk.simpledialog.askstring('Rechnungsnummer',
                                                           "Die letze eingetragen Rechnungsnummer hatte nicht das richtige Format. \nGib sie in dem Format ein wie 2024-001 wobei die erste Nummer mit dem Jahr der Rechnung ersetzt wird und die zweite mit der Rechnungsnummer:",
                                                           parent=root)
                root.destroy()
    return thisinvoicenumber
def validate_input_int(char, input_value):
    """Function to validate input - allows only integer values."""
    # If the input is empty, it's valid (so the user can delete the input).
    if input_value == "":
        return True
    try:
        # Try to convert the input value to an integer
        int(input_value)
        return True
    except ValueError:
        return False
def change_place_of_window(root):
    w = 800  # width for the Tk root
    h = 650  # height for the Tk root

    # get screen width and height
    ws = root.winfo_screenwidth()  # width of the screen
    hs = root.winfo_screenheight()  # height of the screen
    # calculate x and y coordinates for the Tk root window
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)

    # set the dimensions of the screen
    # and where it is placed
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))



def get_selection(title):

    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Prompt the user with a message box
    response = messagebox.askyesno("Eingeben?", title)

    # Return the user's response
    return response




def select_client(options):
    root = tk.Tk()
    root.title('Patientenauswahl:')
    # prompt =  "Von welcher Person willst du die Rechnung ausdrucken?",
    # tk.Label(root, text=prompt).pack()

    search_entry = tk.Entry(root, width=80)
    search_entry.pack(pady=10)

    def filter_list(event):
        searching_for = search_entry.get()
        print(searching_for)
        if isinstance(searching_for,str):
            search_term = search_entry.get().lower()
            filtered_options = [option for option in options if search_term in option.lower()]

            # Clear the current listbox
            listbox.delete(0, tk.END)

            # Add filtered options to the listbox
            for option in filtered_options:
                listbox.insert(tk.END, option)
    search_entry.bind('<KeyRelease>', filter_list)

    frame = tk.Frame(root)
    frame.pack(pady=10)

    # Create a scrollable listbox
    listbox = tk.Listbox(frame, height=15, width=80, selectmode=tk.SINGLE)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH)

    scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=listbox.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Link the scrollbar to the listbox
    listbox.config(yscrollcommand=scrollbar.set)

    # Populate the listbox with all options initially
    for option in options:
        listbox.insert(tk.END, option)

    selected_name = [""]
    def on_ok():
        selected = listbox.curselection()
        if selected:
            index, = listbox.curselection()
            selected_name[0] =  options[index]
            # selected_name = options[index]
            root.destroy()


    tk.Button(text="OK", command=on_ok,font=("Helvetica", 14) ).pack()
    root.mainloop()


    return selected_name[0]


def save_to_archive(invoicenumber,datetoday,clientname,invoice_start_date,invoice_end_date,summe,archive_which_invoices_path):
    ws_archive_which_invoices = openpyxl.load_workbook(archive_which_invoices_path)
    archive_which_invoices = ws_archive_which_invoices.worksheets[0]
    invoiceduration = invoice_start_date.strftime("%d.%m.%Y") + " - " + invoice_end_date.strftime("%d.%m.%Y")
    #datetoday = datetime.datetime.strptime(datetoday, "%d.%m.%Y")
    inputdata = [invoicenumber,datetoday, clientname, invoiceduration, summe]

    last_row = archive_which_invoices.max_row
    last_row_data = 0 #last row of invoice data (not sums etc)
    sumrow = True
    i = last_row
    while True:
        if sumrow:
            if archive_which_invoices.cell(i, 1).value is not None: # check whether there is a sum row
                i -=1
            else:
                sumrow = False
        else:
            if archive_which_invoices.cell(i, 1).value is None:  # check whether there is an empty row over a sum row
                i -= 1
            else:
                print(f"last datarow = {i}, {[cell.value for cell in archive_which_invoices[i]]}")
                last_row_data = i
                break
    archive_which_invoices.insert_rows(last_row_data+1)
    last_row = archive_which_invoices.max_row
    for col, value in zip(range(1, len(inputdata) + 1), inputdata):
        archive_which_invoices.cell(row=last_row_data+1, column=col, value=value)
        archive_which_invoices.cell(last_row_data+1,5).number_format = '€* #,##0.00'
        archive_which_invoices.cell(last_row_data+1,2).number_format = 'DD.MM.YYYY'

    cell_sum_invoiced = archive_which_invoices.cell(last_row-1, 5)
    cell_sum_paid = archive_which_invoices.cell(last_row-1, 7)
    cell_diff_inv_paid = archive_which_invoices.cell(last_row, 7)

    cell_sum_invoiced.value = f"=SUM(E2:E{last_row-2})"
    cell_sum_paid.value = f"=SUM(G2:G{last_row-2})"
    cell_diff_inv_paid.value = f"=E{last_row-1}-G{last_row-1}"



    print(f"Saved the archive excel to: {archive_which_invoices_path}")
    ws_archive_which_invoices.save(archive_which_invoices_path)




def show_matrix_window(matrix, frame, head = ("",""),lastdate = 0,fill='both',expand = True,padx=10, pady=10):
    print(lastdate)
    if lastdate:
        Title = tk.Label(frame, text=f"Die letze Rechnung für diese Person wurde am {lastdate} erstellt").pack(pady=10)
    treeview = ttk.Treeview(frame, columns=head, show="headings")

    for colname in head:
        treeview.heading(colname, text=colname)


    for column in matrix:
        columntupel = tuple(column)
        if isinstance(columntupel[1],pd.DataFrame):
            temp_list = list(columntupel)
            temp_list[1] = columntupel[1].to_string(index=False)
            columntupel = tuple(temp_list)


        treeview.insert("", tk.END,values=columntupel)
    treeview.pack(padx=padx, pady=pady,fill=fill,expand = expand)

    def motion_handler(tree, event):
        f = Font(font='TkDefaultFont')

        # A helper function that will wrap a given value based on column width
        def adjust_newlines(val, width, pad=10):
            if not isinstance(val, str):
                return val
            else:
                words = val.split()
                lines = [[], ]
                for word in words:
                    line = lines[-1] + [word, ]
                    if f.measure(' '.join(line)) < (width - pad):
                        lines[-1].append(word)
                    else:
                        lines[-1] = ' '.join(lines[-1])
                        lines.append([word, ])

                if isinstance(lines[-1], list):
                    lines[-1] = ' '.join(lines[-1])

                return '\n'.join(lines)

        if (event is None) or (tree.identify_region(event.x, event.y) == "separator"):
            # You may be able to use this to only adjust the two columns that you care about
            # print(tree.identify_column(event.x))

            col_widths = [tree.column(cid)['width'] for cid in tree['columns']]

            for iid in tree.get_children():
                new_vals = []
                for (v, w) in zip(tree.item(iid)['values'], col_widths):
                    new_vals.append(adjust_newlines(v, w))
                tree.item(iid, values=new_vals)

    def calculate_row_height(tree):
        """Calculate and adjust row height to fit the text."""
        # Retrieve the existing Treeview font
        style = ttk.Style()
        font = Font(name="TkDefaultFont", exists=True)  # Use the existing font

        # Determine the required height for the tallest text
        max_text_height = 0
        for item in tree.get_children():
            row_values = tree.item(item, "values")
            for text in row_values:
                # Measure the height of the text
                max_text_height = max(max_text_height, font.metrics("linespace"))

        # Adjust the row height dynamically
        style.configure("Treeview", rowheight=max_text_height + 10)  # Add padding

    treeview.bind('<B1-Motion>', partial(motion_handler, treeview))
    motion_handler(treeview, None)   # Perform initial wrapping
    calculate_row_height(treeview)

def ask_to_save(data_list):
    root = tk.Tk()
    root.title("Überprüfung")
    root.geometry("700x800+50+30")
    Label(root, text="Ich erstelle nun eine Rechnung mit folgenden Daten:").pack()
    data_list_without_hours = [x for x in data_list if "Stundeninfo" not in x[0] ]
    show_matrix_window(data_list_without_hours, root, head = ("","Wert"), fill="x",padx=0,pady=0)
    Label(root, text="Stundendaten").pack()
    hourdata =[x for x in data_list if "Stundeninfo" in x[0]][0][1]
    show_matrix_window(list(hourdata.values),root,head=tuple(hourdata.columns),fill="x",padx=0,pady=0)

    Label(root, text="Soll ich nun einen Rechnung mit diesen Daten erstellen?").pack()
    yes_no_frame = tk.Frame(root)
    yes_no_frame.pack()
    ttk.Style().configure('Treeview', rowheight=30)

    answer = [False]
    def button_press(y_n):
        if y_n == "Y":
            print("Y")
            answer[0] = True
        elif y_n == "N":
            print("N")
            answer[0] = False
        root.destroy()
    # Pack widgets side by side inside the last row frame
    tk.Button(yes_no_frame, text="Ja",command=lambda: button_press("Y")).pack(side="left", padx=10)
    tk.Button(yes_no_frame, text="Nein",command=lambda: button_press("N")).pack(side="left", padx=10)

    root.mainloop()
    return answer[0]

def ask_right_invoicenumber(question):
    root = tk.Tk()
    root.title("Frage")

    Label(root, text=question).pack(padx=10, pady=10)
    yes_no_frame = tk.Frame(root)
    yes_no_frame.pack(pady=10,expand=True)
    ttk.Style().configure('Treeview', rowheight=30)

    answer = [True]
    def button_press_y():
        print("Y")
        answer[0] = True
        root.destroy()
    def button_press_n():
        print("N")
        answer[0] = False
        root.destroy()
    # Pack widgets side by side inside the last row frame
    tk.Button(yes_no_frame, text="OK",command=button_press_y).pack(side="left", padx=10)
    tk.Button(yes_no_frame, text="Ändern",command=button_press_y).pack(side="left", padx=10)
    root.mainloop()
    print("proceed")
    print("Window closed, proceeding with the program.")
    return answer[0]

def get_date(hourdata,lastdate):
    selected_date = None
    root = tk.Tk()
    root.title("Auswahl des Rechnungszeitraums")

    left_frame = tk.Frame(root)
    right_frame = tk.Frame(root)

    left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)


    label1 = Label(left_frame, text='Rechnung ab: ', font=("Helvetica", 14) )
    label1.pack(ipadx=10, ipady=10)

    cal1 = Calendar(left_frame,
                   font="Arial 14", selectmode='day')
    cal1.pack(fill="both", expand=True)
    label2 = Label(left_frame, text='bis: ', font=("Helvetica", 14) )
    label2.pack(ipadx=10, ipady=10)
    cal2 = Calendar(left_frame,
                   font="Arial 14", selectmode='day')
    cal2.pack(fill="both", expand=True)

    tk.Button(left_frame, text="ok",height=2, width=20, font="Arial 14", command=root.destroy).pack(pady=10)
    data_list = hourdata.values.tolist()
    show_matrix_window(data_list, right_frame,head = ("Rechnungsdatum","PatientIn","Stundendauer in min")  , lastdate=lastdate )
    root.mainloop()
    return cal1.selection_get(), cal2.selection_get()


def input_new_person(allclientdata_path):
    allclientdata = pd.read_excel(allclientdata_path, index_col=0, header=None, sheet_name=None)

    datatoinquire = list(allclientdata["Vorlage"].index)
    #

    root = Tk()
    # initialise the boxes
    labels = [Label(root, text = onedatalabel) for onedatalabel in datatoinquire]
    entries = [Entry(root) for x in range(0,len(datatoinquire))]


    #position the inquiries in a nice table
    for rownumber, (label, entry) in enumerate(zip(labels, entries)):
        label.grid(column=0, row=rownumber)
        if rownumber != 1 and rownumber != 2:
            entry.grid(column=1, row=rownumber)

    #make the dropdowns
    sexoptions = ["w","m"]
    childoptions =["ja", "nein"]


    child = StringVar(root)
    child.set(childoptions[0])
    childoptiondropdown = OptionMenu(root, child, *childoptions)
    childoptiondropdown.grid(column=1, row=1)

    sex = StringVar(root)
    sex.set(sexoptions[0])
    sexoptiondropdown = OptionMenu(root, sex, *sexoptions)
    sexoptiondropdown.grid(column=1, row=2)


    userinputs = []
    def command():
        for entry in entries:
            userinputs.append(entry.get())
        userinputs[1] = child.get()
        userinputs[2] = sex.get()
        root.destroy()

    Button(root, text="Speichern", command=command).grid(column=1,row =len(datatoinquire)+1)
    root.mainloop()
    # parse datetime inputs
    try:
        userinputs[3] = dateutil.parser.parse(userinputs[3])
    except:
        print("Das Datum ist falsch eingegeben")

    try:
        userinputs[12] = dateutil.parser.parse(userinputs[12])
    finally:

        userinputsdict = dict(zip(datatoinquire, userinputs))


        excelsheet_with_added_person = openpyxl.load_workbook(allclientdata_path)#
        excelsheet_with_added_person.iso_dates = True
        sheet_new_person = excelsheet_with_added_person.create_sheet(userinputsdict["Name"])


        for row, (dataname,userinput) in enumerate(zip(datatoinquire,userinputs)):
            sheet_new_person.cell(row=row+1, column=1).value = dataname
            sheet_new_person.cell(row=row+1, column=2).value = userinput
        excelsheet_with_added_person.save(allclientdata_path)

        print("Ich habe eine neues Blatt für " + userinputsdict["Name"] + " zur PatienInneninformations Exceldatei hinzugefügt")
        print("Mit diesen Einträgen: ")
        pprint.pprint(userinputsdict)
        return userinputsdict


def insert_hourdata(allhourdata_path,clientname):
    root = Tk()
    root.title("Therapiedaten für " + clientname)
    root.geometry("650x500+120+120")

    # empty arrays for your Entrys and StringVars
    text_var = []
    entries = []

    # callback function to get your StringVars
    clienthourdata = []
    def command():
        matrix = []
        for i in range(rows):
            matrix.append([])
            for j in range(cols):
                matrix[i].append(text_var[i][j].get())
        clienthourdata.append(matrix)
        root.destroy()

    labelnames = ["Datum Therapie (im Format wie 1.1.2023)", "Einheitslänge in min"]
    for column in range(0,2):
        Label(root, text=labelnames[column], font=('arial', 10, 'bold'),
          bg="bisque2").place(x=20 + 110*column, y=20)

    x2 = 0
    y2 = 0
    rows, cols = (10,2)
    for i in range(rows):
        # append an empty list to your two arrays
        # so you can append to those later
        text_var.append([])
        entries.append([])
        for j in range(cols):
            # append your StringVar and Entry
            text_var[i].append(StringVar())
            entries[i].append(Entry(root, textvariable=text_var[i][j],width=10))
            entries[i][j].place(x=60 + x2, y=50 + y2)
            x2 += 100

        y2 += 30
        x2 = 0
    button= Button(root,text="Daten speichern", bg='bisque3', width=15, command=command)
    button.place(x=160,y=350)
    root.mainloop()


    clienthourdata = clienthourdata[0]
    datestherapy = list(filter(None,[row[0] for row in clienthourdata] ))
    lengththerapy = list(filter(None,[row[1] for row in clienthourdata]))
    print("unparsed")
    print(datestherapy)
    datestherapy = list(map(lambda x: dateutil.parser.parse(x, dayfirst = True), datestherapy))
    lengththerapy = list(map(lambda x: float(x), lengththerapy))
    print("parsed")
    print(datestherapy)

    excelsheet_hourdata = openpyxl.load_workbook(allhourdata_path)  #
    excelsheet_hourdata.iso_dates = True
    sheet = excelsheet_hourdata["Stundendaten"]

    for dateonetherapy, lengthonetherapy in zip(datestherapy, lengththerapy):
        newRowLocation = sheet.max_row + 1
        sheet.cell(row=newRowLocation, column=1).value = dateonetherapy
        sheet.cell(row=newRowLocation, column=2).value = clientname
        sheet.cell(row=newRowLocation, column=3).value = lengthonetherapy


    excelsheet_hourdata.save(allhourdata_path)
    namehourdata = pd.DataFrame([datestherapy,[clientname for x in range(0,len(datestherapy))],lengththerapy])
    namehourdata = namehourdata.transpose()
    namehourdata.columns = ['Datum', 'Name', 'Minuten']

    return(namehourdata)


# start,end = get_date()