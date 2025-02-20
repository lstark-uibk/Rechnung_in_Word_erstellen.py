import os
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import dateutil.parser
import openpyxl
import xlwings as xw


def make_new_Person(allclientdata_path):

    allclientdata = pd.read_excel(allclientdata_path, index_col=0, header=None, sheet_name=None)

    datatoinquire = list(allclientdata["Vorlage"].index)
    #

    root = tk.Tk()
    root.title("Lege eine neue Person an")
    # initialise the boxes
    labels = [tk.Label(root, text=onedatalabel) for onedatalabel in datatoinquire]
    entries = [tk.Entry(root) for x in range(0, len(datatoinquire))]

    # position the inquiries in a nice table
    for rownumber, (label, entry) in enumerate(zip(labels, entries)):
        label.grid(column=0, row=rownumber)
        if rownumber != 1 and rownumber != 2:
            entry.grid(column=1, row=rownumber)

    # make the dropdowns
    sexoptions = ["w", "m"]
    childoptions = ["ja", "nein"]

    child = tk.StringVar(root)
    child.set(childoptions[0])
    childoptiondropdown = tk.OptionMenu(root, child, *childoptions)
    childoptiondropdown.grid(column=1, row=1)

    sex = tk.StringVar(root)
    sex.set(sexoptions[0])
    sexoptiondropdown = tk.OptionMenu(root, sex, *sexoptions)
    sexoptiondropdown.grid(column=1, row=2)

    userinputs = [""]*len(entries)

    def validate_date(x):
        if x:
            try:
                date = dateutil.parser.parse(x)
                return ("",True)
            except:
                return ("Ein Datum ist falsch eingegeben",False)
        else: return ("",True)
    def validate_mandatory(x):
        if x:
            return ("",True)
        else: return ("Das Namensfeld darf nicht leer sein",False)

    def submit():
        # collect the values
        Errormessages.config(text="")
        for i,entry in enumerate(entries):
            userinputs[i] = entry.get()
            entry.config(bg="white")

        userinputs[1] = child.get()
        userinputs[2] = sex.get()
        print(userinputs)
        validated_wrong = []
        errormessages = []
        #validate mandatory
        mandatoryentries = [0]
        errormessage, validation =validate_mandatory(userinputs[0])
        if not validation:
            validated_wrong.append(0)
            errormessages.append(errormessage)
        #validate datetime:
        dateentries = [3,12,16]
        for i in dateentries:
            errormessage, validation = validate_date(userinputs[i])
            if not validation:
                validated_wrong.append(i)
                errormessages.append(errormessage)
        # if nothing was validated wrong proceed
        if not validated_wrong:
            print("Alles Richtig eingegeben")
            for i in dateentries:
                try:
                    userinputs[i] = dateutil.parser.parse(userinputs[i])
                except:
                    pass

            userinputsdict = dict(zip(datatoinquire, userinputs))
            print()

            excelsheet_with_added_person = openpyxl.load_workbook(allclientdata_path)
            excelsheet_with_added_person.iso_dates = True
            sheet_new_person = excelsheet_with_added_person.create_sheet(userinputsdict["Name"])

            for row, (dataname, userinput) in enumerate(zip(datatoinquire, userinputs)):
                sheet_new_person.cell(row=row + 1, column=1).value = dataname
                sheet_new_person.cell(row=row + 1, column=2).value = userinput
            try:
                excelsheet_with_added_person.save(allclientdata_path)
            except: messagebox.showinfo("Fehler","Konnte die Excel nicht speichern.\nSind PatientInneninfomationen schon in excel geöffnet? \nWenn ja schließe diese bitte")
            if os.name == 'posix':
                print("This system is Linux or another Unix-like system.")
                import subprocess
                from sys import platform
                if platform == 'darwin':  # apple
                    subprocess.call(['open', allclientdata_path])

                else:
                    subprocess.run(['xdg-open', allclientdata_path])

            else:
                print("This system is not Linux.")
                os.startfile(allclientdata_path)
    Errormessages = tk.Label(root, textvariable="")
    Errormessages.grid(column=1, row=len(datatoinquire) + 2)
    tk.Button(root, text="Speichern", command=submit).grid(column=1, row=len(datatoinquire) + 1)

    root.mainloop()
# parse datetime inputs
